import telebot
from telebot.types import InlineKeyboardButton as ikb
from telebot.types import KeyboardButton as rkb
from telebot.types import InlineKeyboardMarkup as ikm
from telebot.types import *
import openpyxl
from time import sleep
import json
import os
from threading import Thread
path = os.path.dirname(os.path.abspath(__file__))
src = path + '/config.xlsx'
tb = openpyxl.load_workbook(src)
hobbys = [i.value for i in tb['Лист1']['A']]
regions = [i.value for i in tb['Лист2']['A']]
regkb = ReplyKeyboardMarkup()
for i in regions:
	regkb.add(KeyboardButton(i))
src = path + '/settings.xlsx'
tb = openpyxl.load_workbook(src)
token = tb['Лист1']['A1'].value
provtoken = tb['Лист1']['A2'].value
admin = int(tb['Лист1']['A3'].value)
faqtext = tb['Лист1']['A4'].value
mainchat = tb['Лист1']['A5'].value
evchat = tb['Лист1']['A6'].value
prices = [LabeledPrice(label='Пропуск на мероприятие', amount=50000)]
with open(path + '/data.json') as file:
	js = json.load(file)
answuser = 0
partners = js[1].copy()
users = js[0].copy()
amd = js[2].copy()
ptd = js[3].copy()
usd = js[4].copy()
allevents = js[5].copy()
startkb = ReplyKeyboardMarkup().add(ikb('Мой профиль')).add(rkb('Общение')).add(rkb('Мероприятия')).add(rkb('Помощь'))
adminkb = ikm().add(ikb('Мои мероприятия', callback_data='myevents')).add(ikb('Загрузить списки', callback_data='loadlists')).add(ikb('Отчет по пользователям', callback_data='userlog')).add(ikb('Отчет по мероприятиям', callback_data='eventlog'))
bot = telebot.TeleBot(token)
@bot.message_handler(content_types=['text'])
def start(msg : Message):
	user = msg.from_user.id
	if msg.text == '/start':
		if user == admin:
			bot.send_message(msg.from_user.id,'Главное меню',reply_markup=adminkb)
		elif user in users:
			bot.send_message(msg.from_user.id,'Главное меню',reply_markup=startkb)
		else:
			bot.send_message(user, 'Что может этот бот: Приветствую тебя на платформе здоровья',reply_markup=ikm().add(ikb('Запустить бота',callback_data='reg')))
	elif msg.text == 'Мой профиль':
		profile = usd[str(user)]
		string = 'Имя: ' + profile['name']
		if profile['surname'] != 'Пропустить':
			string += '\nФамилия: ' + profile['surname']
		if profile['sex'] != 'Пропустить':
			string += '\nПол: ' + profile['sex']
		string += '\nВозраст: ' + profile['age']
		string += '\nРегион: ' + profile['region']
		prhobbys = []
		for i in profile['hobbys']:
			if profile['hobbys'][i] == True:
				prhobbys.append(i)
		string += '\nУвлечения: ' + ', '.join(prhobbys)
		kb = ikm()
		if not user in partners:
			kb.add(ikb('Подать заявку на Партнера', callback_data='newpart'))
		if profile['photo'] != 'Пропустить':
			bot.send_photo(user, profile['photo'], string,reply_markup=kb)
		else:
			bot.send_message(user, string,reply_markup=kb)
	elif msg.text == 'Общение':
		bot.send_message(user, 'Наши чаты', reply_markup=ikm().add(ikb('Чат', url=mainchat)).add(ikb('Группа мероприятий', url=evchat)))
	elif msg.text == 'Мероприятия':
		kb = ikm()
		if user in partners:
			kb.add(ikb('Мои мероприятия', callback_data='myevents'))
		kb.add(ikb('Мероприятия Мастера', callback_data='mastevents'))
		kb.add(ikb('Мероприятия Партнеров', callback_data='partevents'))
		bot.send_message(user, 'Меню мероприятий:', reply_markup=kb)
	elif msg.text == 'Помощь':
		bot.send_message(user, faqtext, reply_markup=ikm().add(ikb('Связаться с нами',callback_data='question')))
@bot.callback_query_handler(func=lambda call:True)
def callback(call : CallbackQuery):
	user = call.from_user.id
	if call.data == 'reg' and not user in users:
		usd[str(user)] = {'hobbys' : {}}
		for i in hobbys:
			usd[str(user)]['hobbys'][i] = False
		bot.register_next_step_handler(bot.edit_message_text('Введите ваше имя:', call.message.chat.id, call.message.message_id), regname)
	elif call.data == 'newpart':
		partners.append(user)
		ptd[str(user)] = []
		bot.register_next_step_handler(bot.send_message(user, 'Напишите заявку на партнера. В ней укажите каким родом услуг вы занимаетесь, а так же причину по которой вы хотели бы стать партнером. Также к сообщению прикрепите свое фото.', reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Отменить'))), newpart)
	elif call.data == 'myevents':
		if user == admin and amd == []:
			bot.send_message(user,'Ваш список мероприятий пуст!', reply_markup=ikm().add(ikb('Добавить мероприятие', callback_data='newevent')))
		elif user != admin and ptd[str(user)] == []:
			bot.send_message(user,'Ваш список мероприятий пуст!', reply_markup=ikm().add(ikb('Добавить мероприятие', callback_data='newevent')))
		else:
			if user == admin:
				event = amd[0]
			else:
				event = ptd[str(call.from_user.id)][0]
			kb = ikm().add(ikb('-->', callback_data=f'myevent{len(ptd[str(call.from_user.id)] if user != admin else amd) * -1 + 1}')).add(ikb(f'1/{len(ptd[str(call.from_user.id)] if user != admin else amd)}',callback_data='[]')).add(ikb('Добавить мероприятие',callback_data='newevent'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			if event['photo'] != 'Пропустить':
				bot.send_photo(user, event['photo'], string,reply_markup=kb)
			else:
				bot.send_message(user, string,reply_markup=kb)
	elif call.data.startswith('myevent'):
		data = int(call.data[7:])
		if user == admin and len(amd) < data * -1 or data == 0:
			bot.send_message(user, 'Это конец списка!')
		elif len(ptd[str(call.from_user.id)]) < data * -1 or data == 0:
			bot.send_message(user, 'Это конец списка!')
		else:
			if user == admin:
				event = amd[data]
			else:
				event = ptd[str(call.from_user.id)][data]
			kb = ikm().row(ikb('<--', callback_data=f'myevent{data - 1}'),ikb('-->', callback_data=f'myevent{data + 1}')).add(ikb(f'{len(ptd[str(call.from_user.id)] if user != admin else amd) + data + 1}/{len(ptd[str(call.from_user.id)] if user != admin else amd)}',callback_data='[]')).add(ikb('Добавить мероприятие',callback_data='newevent'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			bot.delete_message(user, call.message.id)
			if event['photo'] != 'Пропустить':
				bot.send_photo(user,event['photo'], string, reply_markup=kb)
			else:
				bot.send_message(user,string,reply_markup=kb)
	elif call.data == 'newevent':
		if call.from_user.id != admin:
			ptd[str(user)].insert(0, {'isready' : False,'hobbys':{},'members' : []})
			for i in hobbys:
				ptd[str(user)][0]['hobbys'][i] = False
			bot.register_next_step_handler(bot.send_message(user,'Введите название мероприятия:'), evname)
		else:
			amd.insert(0, {'isready' : False,'hobbys':{},'members' : []})
			for i in hobbys:
				amd[0]['hobbys'][i] = False
			bot.register_next_step_handler(bot.send_message(user,'Введите название мероприятия:'), evname)
	elif call.data == 'partevents':
		if len(allevents) != 0:
			event = ptd[str(allevents[0][0])][allevents[0][1]]
			kb = ikm().add(ikb('-->', callback_data=f'pevent{len(allevents) * -1 + 1}')).add(ikb(f'1/{len(allevents)}',callback_data='[]')).add(ikb('Учавствовать',callback_data=f'inv{len(allevents) * -1}'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			if event['photo'] != 'Пропустить':
				bot.send_photo(user, event['photo'], string,reply_markup=kb)
			else:
				bot.send_message(user, string,reply_markup=kb)
		else:
			bot.send_message(user,'Пока что никто не создал ни одного мероприятия!')
	elif call.data.startswith('pevent'):
		data = int(call.data[6:])
		if len(allevents) < data * -1 or data == 0:
			bot.send_message(user, 'Это конец списка!')
		else:
			event = ptd[str(allevents[data][0])][allevents[data][1]]
			kb = ikm().row(ikb('<--', callback_data=f'pevent{data - 1}'),ikb('-->', callback_data=f'pevent{data + 1}')).add(ikb(f'{len(allevents) + data + 1}/{len(allevents)}',callback_data='[]')).add(ikb('Учавствовать',callback_data=f'inv{data}'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			bot.delete_message(user, call.message.id)
			if event['photo'] != 'Пропустить':
				bot.send_photo(user,event['photo'], string, reply_markup=kb)
			else:
				bot.send_message(user,string,reply_markup=kb)
	elif call.data == 'mastevents':
		if len(amd) != 0:
			event = amd[0]
			kb = ikm().add(ikb('-->', callback_data=f'mevent{len(amd) * -1 + 1}')).add(ikb(f'1/{len(amd)}',callback_data='[]')).add(ikb('Учавствовать',callback_data=f'minv{len(amd) * -1}'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			if event['photo'] != 'Пропустить':
				bot.send_photo(user, event['photo'], string,reply_markup=kb)
			else:
				bot.send_message(user, string,reply_markup=kb)
		else:
			bot.send_message(user,'Пока что Мастер не создал ни одного мероприятия!')
	elif call.data.startswith('mevent'):
		data = int(call.data[6:])
		if len(amd) < data * -1 or data == 0:
			bot.send_message(user, 'Это конец списка!')
		else:
			event = amd[data]
			kb = ikm().row(ikb('<--', callback_data=f'mevent{data - 1}'),ikb('-->', callback_data=f'mevent{data + 1}')).add(ikb(f'{len(amd) + data + 1}/{len(amd)}',callback_data='[]')).add(ikb('Учавствовать',callback_data=f'minv{data}'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			bot.delete_message(user, call.message.id)
			if event['photo'] != 'Пропустить':
				bot.send_photo(user,event['photo'], string, reply_markup=kb)
			else:
				bot.send_message(user,string,reply_markup=kb)
	elif call.data == 'loadlists':
		bot.register_next_step_handler(bot.send_message(user,'Отправьте таблицу с данными:'), loadlist)
	elif call.data.startswith('inv'):
		bot.send_invoice(call.from_user.id, 'Пропуск на мероприятие', 'Данный пропуск позволяет вам пройти на "' + ptd[allevents[int(call.data[3:])][0]][allevents[int(call.data[3:])][1]]['name'] + '".', f'p{call.data[3:]}', provtoken,'rub', [LabeledPrice(label='Пропуск на мероеприятие', amount=ptd[allevents[int(call.data[3:])][0]][allevents[int(call.data[3:])][1]]['price'])], start_parameter='invite')
	elif call.data.startswith('minv'):
		bot.send_invoice(call.from_user.id, 'Пропуск на мероприятие', 'Данный пропуск позволяет вам пройти на "' + amd[int(call.data[4:])]['name'] + '".', f'm{call.data[4:]}', provtoken,'rub', [LabeledPrice(label='Пропуск на мероеприятие', amount=amd[int(call.data[4:])]['price'])], start_parameter='invite')
	elif call.data == 'userlog':
		wb = openpyxl.Workbook()
		ws = wb.active
		for i in usd:
			try:
				print([usd[i]['name'],usd[i]['surname'] if usd[i]['surname'] != 'Пропустить' else '',usd[i]['sex'] if usd[i]['sex'] != 'Пропустить' else '',usd[i]['age'],usd[i]['region'],', '.join(usd[i]['hobbys']),usd[i]['photo'] if usd[i]['photo'] != 'Пропустить' else ''])
				ws.append([usd[i]['name'],usd[i]['surname'] if usd[i]['surname'] != 'Пропустить' else '',usd[i]['sex'] if usd[i]['sex'] != 'Пропустить' else '',usd[i]['age'],usd[i]['region'],', '.join(usd[i]['hobbys']),usd[i]['photo'] if usd[i]['photo'] != 'Пропустить' else ''])
			except:
				pass
		wb.save(path + '/members_log.xlsx')
		bot.send_document(call.from_user.id, InputFile(path + '/members_log.xlsx'))
	elif call.data == 'eventlog':
		wb = openpyxl.Workbook()
		ws = wb.active
		for i in allevents:
			try:
				event = ptd[i[0]][i[1]]
				print([event['name'], event['date'], event['time'], ', '.join(event['hobbys']), event['desc'] if event['desc'] != 'Пропустить' else '', event['photo'] if event['photo'] != 'Пропустить' else ''])
				ws.append([event['name'], event['date'], event['time'], ', '.join(event['hobbys']), event['desc'] if event['desc'] != 'Пропустить' else '', event['photo'] if event['photo'] != 'Пропустить' else ''])
			except:
				pass
		wb.save(path + '/events_log.xlsx')
		bot.send_document(call.from_user.id, InputFile(path + '/events_log.xlsx'))
	elif call.data.startswith('npart'):
		part = int(call.data[5:])
		partners.append(part)
		ptd[str(part)] = []
		bot.send_message(part, 'Вашу заявку на партнера одобрили!')
	elif call.data.startswith('cpart'):
		part = int(call.data[5:])
		bot.send_message(part, 'Вашу заявку на партнера отклонили!')
	elif call.data == 'question':
		bot.register_next_step_handler(bot.send_message(user,'Напишите свой вопрос и мы постараемся вам на него ответить!'), question)
	elif call.data.startswith('answer'):
		global answuser
		answuser = int(call.data[6:])
		bot.register_next_step_handler(bot.send_message(admin,'Введите ответ на вопрос пользователя:'), answer)
@bot.pre_checkout_query_handler(func=lambda query: True)
def checkout(pre_checkout_query : PreCheckoutQuery):
    bot.answer_pre_checkout_query(pre_checkout_query.id, ok=True,error_message="Что-то пошло не так! Повторите попытку")
@bot.message_handler(content_types=['successful_payment'])
def got_payment(message : Invoice):
	bot.send_message(message.from_user.id, 'Проход на мероприятие успешно приобретён!')
	data = int(message.successful_payment.invoice_payload[1:])
	if message.successful_payment.invoice_payload[0] == 'p':
		ptd[allevents[data][0]][allevents[data][1]]['members'].append(message.from_user.id)
	else:
		amd[data]['members'].append(message.from_user.id)
def question(msg : Message):
	if msg.content_type == 'text':
		bot.send_message(admin, f'Пользователь хочет задать вам вопрос: {msg.text}', reply_markup=ikm().add(ikb('Ответить',callback_data=f'answer{msg.from_user.id}')))
		bot.send_message(msg.from_user.id, 'Ваша заявка отправленна модерации!')
	else:
		bot.send_message(msg.from_user.id, 'Пожалуйста отправьте текст!')
def answer(msg : Message):
	if msg.content_type == 'text':
		bot.send_message(answuser, f'Модерация ответила на ваш вопрос: {msg.text}')
		bot.send_message(answuser, 'Ответ отправлен пользователю')
	else:
		bot.send_message(msg.from_user.id, 'Пожалуйста отправьте текст!')
def newpart(msg : Message):
	if msg.content_type == 'text' and msg.text == 'Отменить':
		bot.send_message(msg.from_user.id, 'Подача заявки отменена.')
	if msg.content_type != 'photo':
		bot.send_message(msg.from_user.id,'Пожалуйста отправьте заявку с прикрепленной фотографией:')
	else:
		bot.send_photo(admin, msg.photo[len(msg.photo) - 1].file_id, usd[str(msg.from_user.id)]['name'] + ' ' + usd[str(msg.from_user.id)]['surname'] + f' подал заявку на партнера. Текст заявки: {msg.caption}', reply_markup=ikm().add(ikb('Принять', callback_data=f'npart{msg.from_user.id}')).add(ikb('Отклонить', callback_data=f'cpart{msg.from_user.id}')))
		bot.send_message(msg.from_user.id, 'Ваша заявка отправленна администрации. Ожидайте!')
def loadlist(msg : Message):
	global regions, hobbys, regkb
	if msg.content_type == 'document':
		file_info = bot.get_file(msg.document.file_id)
		downloaded_file = bot.download_file(file_info.file_path)
		src = path + '/config.xlsx'
		with open(src, 'wb') as new_file:	
			new_file.write(downloaded_file)
		tb = openpyxl.load_workbook(src)
		hobbys = [i.value for i in tb['Лист1']['A']]
		regions = [i.value for i in tb['Лист2']['A']]
		regkb = ReplyKeyboardMarkup()
		for i in regions:
			regkb.add(KeyboardButton(i))
		bot.send_message(msg.from_user.id,'Успешно!')
	else:
		bot.send_message(msg.from_user.id,'Это не похоже на таблицу!')
def regname(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Это не похоже на имя! Введите ваше имя:'), regname)
	elif not any(character.isdigit() for character in msg.text):
		usd[str(msg.from_user.id)]['name'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Введите вашу фамилию(не обязательно):', reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), regsurname)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'В имени не может быть цифр! Введите ваше имя:'), regname)
def regsurname(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Это не похоже на фамилию! Введите вашу фамилию:', reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), regsurname)
	elif not any(character.isdigit() for character in msg.text):
		usd[str(msg.from_user.id)]['surname'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваш пол:',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Мужской'),KeyboardButton('Женский')).add(KeyboardButton('Пропустить'))), regsex)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'В фамилии не может быть цифр! Введите вашу фамилию:', reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), regsurname)
def regsex(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Это не похоже на пол, для выбора пола воспользуйтесь кнопками на клавиатуре',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Мужской'),KeyboardButton('Женский')).add(KeyboardButton('Пропустить'))), regsex)
	elif msg.text == 'Мужской' or msg.text == 'Женский' or msg.text == 'Пропустить':
		usd[str(msg.from_user.id)]['sex'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Напишите сколько вам лет:', reply_markup=ReplyKeyboardRemove()), regage)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Это не похоже на пол, для выбора пола воспользуйтесь кнопками на клавиатуре',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Мужской'),KeyboardButton('Женский')).add(KeyboardButton('Пропустить'))), regsex)
def regage(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Возраст должен быть числом! Напишите сколько вам лет:', reply_markup=ReplyKeyboardRemove()), regage)
	elif msg.text.isdigit():
		usd[str(msg.from_user.id)]['age'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваш регион:',reply_markup=regkb), regreg)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Возраст должен быть числом! Напишите сколько вам лет:', reply_markup=ReplyKeyboardRemove()), regage)
def regreg(msg : Message):
	if msg.text in regions and msg.content_type == 'text':
		keys = list(usd[str(msg.from_user.id)]['hobbys'].keys())
		usd[str(msg.from_user.id)]['region'] = msg.text
		kb = ReplyKeyboardMarkup()
		for i in keys:
			if usd[str(msg.from_user.id)]['hobbys'][i] == True:
				kb.add(KeyboardButton(f'{i} \U00002705'))
			else:
				kb.add(KeyboardButton(f'{i} \U0000274C'))
		kb.add(KeyboardButton('Продолжить'))
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваши интересы из списка:',reply_markup=kb), reghobby)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Это не похоже на регион из списка. Выберите регион используя кнопки на клавиатуре', reply_markup=regkb), regreg)
def reghobby(msg : Message):
	if msg.content_type == 'text':
		hobby = msg.text[:-2]
		keys = list(usd[str(msg.from_user.id)]['hobbys'].keys())
		a = 0
		if msg.text == 'Продолжить':
			for i in keys:
				if usd[str(msg.from_user.id)]['hobbys'][i] == True:
					a += 1
			if a > 0:
				bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Загрузите ваше фото:', reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), regphoto)
			else:
				bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите хотя бы 1 пункт'), reghobby)
		elif hobby in keys:
			usd[str(msg.from_user.id)]['hobbys'][hobby] = not usd[str(msg.from_user.id)]['hobbys'][hobby]
			kb = ReplyKeyboardMarkup()
			for i in keys:
				if usd[str(msg.from_user.id)]['hobbys'][i] == True:
					kb.add(KeyboardButton(f'{i} \U00002705'))
				else:
					kb.add(KeyboardButton(f'{i} \U0000274C'))
			kb.add(KeyboardButton('Продолжить'))
			bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Список изменен! Для подтвержднения выбора нажмите "Продолжить".', reply_markup=kb), reghobby)
		else:
			bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваши интересы используя кнокпи на клавиатуре!'), reghobby)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваши интересы используя кнокпи на клавиатуре!'), reghobby)

def regphoto(msg : Message):
	if msg.content_type == 'text' and msg.text == 'Пропустить':
		usd[str(msg.from_user.id)]['photo'] = 'Пропустить'
		users.append(msg.from_user.id)
		bot.send_message(msg.from_user.id,'Регистрация прошла успешно',reply_markup=ReplyKeyboardRemove())
		bot.send_message(msg.from_user.id,'Главное меню',reply_markup=startkb)
	elif msg.content_type == 'photo':
		usd[str(msg.from_user.id)]['photo'] = msg.photo[len(msg.photo) - 1].file_id
		users.append(msg.from_user.id)
		print(usd[str(msg.from_user.id)])
		bot.send_message(msg.from_user.id,'Регистрация прошла успешно',reply_markup=ReplyKeyboardRemove())
		bot.send_message(msg.from_user.id,'Главное меню',reply_markup=startkb)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Это не похоже на фотографию. Загрузите ваше фото:', reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), regphoto)
def evname(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Это не похоже на название! Введите название мероприятия:'), evname)
	else:
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['name'] = msg.text
		else:
			amd[0]['name'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Введите дату начала мероприятия в формате дд.мм.гггг:'), evdate)
def evdate(msg : Message):
	if msg.content_type != 'text' or not msg.text[:1].isdigit() or not msg.text[3:4].isdigit() or not msg.text[6:9].isdigit() or not msg.text[2] == '.' or not msg.text[5] == '.':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Это не похоже на дату! Введите дату начала мероприятия в формате дд.мм.гггг:'), evdate)
	else:
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['date'] = msg.text
		else:
			amd[0]['date'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Введите время начала мероприятия в формате чч:мм:'), evtime)

def evtime(msg : Message):
	if msg.content_type != 'text' or not msg.text[:1].isdigit() or not msg.text[3:4].isdigit() or not msg.text[2] == ':':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Это не похоже на время! Введите время начала мероприятия в формате чч:мм:'), evtime)
	else:
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['time'] = msg.text
		else:
			amd[0]['time'] = msg.text
		kb = ReplyKeyboardMarkup()
		keys = list(ptd[str(msg.from_user.id)][0]['hobbys'].keys() if msg.from_user.id != admin else amd[0]['hobbys'].keys())
		for i in keys:
			if (ptd[str(msg.from_user.id)][0]['hobbys'][i] if msg.from_user.id != admin else amd[0]['hobbys'][i]) == True:
				kb.add(KeyboardButton(f'{i} \U00002705'))
			else:
				kb.add(KeyboardButton(f'{i} \U0000274C'))
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Выберите интересы, соответствующие мероприятию:', reply_markup=kb), evhobby)
def evhobby(msg : Message):
	if msg.content_type == 'text':
		hobby = msg.text[:-2]
		keys = list(ptd[str(msg.from_user.id)][0]['hobbys'].keys() if msg.from_user.id != admin else amd[0]['hobbys'].keys())
		a = 0
		if msg.text == 'Продолжить':
			for i in keys:
				if (ptd[str(msg.from_user.id)][0]['hobbys'][i] if msg.from_user.id != admin else amd[0]['hobbys'][i]) == True:
					a += 1
			if a > 0:
				bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Введите адрес проведения мероприятия:',reply_markup=ReplyKeyboardMarkup()), evadress)
			else:
				bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите хотя бы 1 пункт'), evhobby)
		elif hobby in keys:
			if msg.from_user.id != admin:
				ptd[str(msg.from_user.id)][0]['hobbys'][hobby] = not ptd[str(msg.from_user.id)][0]['hobbys'][hobby]
			else:
				amd[0]['hobbys'][hobby] = not amd[0]['hobbys'][hobby]
			kb = ReplyKeyboardMarkup()
			for i in keys:
				if (ptd[str(msg.from_user.id)][0]['hobbys'][i] if msg.from_user.id != admin else amd[0]['hobbys'][i]) == True:
					kb.add(KeyboardButton(f'{i} \U00002705'))
				else:
					kb.add(KeyboardButton(f'{i} \U0000274C'))
			kb.add(KeyboardButton('Продолжить'))
			bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Список изменен! Для подтвержднения выбора нажмите "Продолжить".', reply_markup=kb), evhobby)
		else:
			bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваши интересы используя кнокпи на клавиатуре!'), evhobby)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id, 'Выберите ваши интересы используя кнокпи на клавиатуре!'), evhobby)

def evadress(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Это не похоже на адрес! Введите адрес проведения мероприятия:'), evadress)
	else:
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['adress'] = msg.text
		else:
			amd[0]['adress'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Введите описание мероприятия:',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), evdesc)
def evdesc(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Это не похоже на описание! Введите описание мероприятия:'), evdesc)
	else:
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['desc'] = msg.text
		else:
			amd[0]['desc'] = msg.text
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Прикрепите фотографию к мероприятию:',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Пропустить'))), evphoto)
def evphoto(msg : Message):
	if msg.content_type == 'text' and msg.text == 'Пропустить':
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['photo'] = 'Пропустить'
		else:
			amd[0]['photo'] = 'Пропустить'
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Введите стоимость входа на мероприятие в рублях:',reply_markup=ReplyKeyboardRemove()), evprice)
	elif msg.content_type != 'photo':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Это не похоже на фотографию! Прикрепите фоторграфию к мероприятию:'), evphoto)
	else:
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['photo'] = msg.photo[len(msg.photo) - 1].file_id
		else:
			amd[0]['photo'] = msg.photo[len(msg.photo) - 1].file_id
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Введите стоимость входа на мероприятие в рублях:',reply_markup=ReplyKeyboardRemove()), evprice)
def evprice(msg : Message):
	if msg.content_type == 'text' and msg.text.isdigit():
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['price'] = int(msg.text) * 100
		else:
			amd[0]['price'] = int(msg.text) * 100
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Сохранить мероприятие?',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Да')).add(KeyboardButton('Нет'))), evconf)
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Цена должна быть числом! Введите стоимость входа на мероприятие в рублях:'), evprice)
def evconf(msg : Message):
	if msg.content_type != 'text':
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Выберите ответ используя кнопки на клавиатуре! Сохранить мероприятие?',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Да').add(KeyboardButton('Нет')))), evconf)
	elif msg.text.lower() == 'да':
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)][0]['ready'] = True
			allevents.insert(0, [str(msg.from_user.id), len(ptd[str(msg.from_user.id)]) * -1])
		else:
			amd[0]['ready'] = True
		bot.send_message(msg.from_user.id, 'Успешно!', reply_markup=ReplyKeyboardRemove())
		if msg.from_user.id == admin and amd == []:
			bot.send_message(msg.from_user.id,'Ваш список мероприятий пуст!', reply_markup=ikm().add(ikb('Добавить мероприятие', callback_data='newevent')))
		elif msg.from_user.id != admin and ptd[str(msg.from_user.id)] == []:
			bot.send_message(msg.from_user.id,'Ваш список мероприятий пуст!', reply_markup=ikm().add(ikb('Добавить мероприятие', callback_data='newevent')))
		else:
			if msg.from_user.id == admin:
				event = amd[0]
			else:
				event = ptd[str(msg.from_user.id)][0]
			kb = ikm().add(ikb('-->', callback_data=f'myevent{len(ptd[str(msg.from_user.id)] if msg.from_user.id != admin else amd) * -1 + 1}')).add(ikb(f'1/{len(ptd[str(msg.from_user.id)] if msg.from_user.id != admin else amd)}',callback_data='[]')).add(ikb('Добавить мероприятие',callback_data='newevent'))
			string = event['name'] + '\nДата начала: ' + event['date'] + '\nВремя начала: ' + event['time']
			prhobbys = []
			for i in event['hobbys']:
				if event['hobbys'][i] == True:
					prhobbys.append(i)
			string += '\nУвлечения: ' + ', '.join(prhobbys)
			if event['desc'] != 'Пропустить':
				string += '\n'
				string += event['desc']
			if event['photo'] != 'Пропустить':
				bot.send_photo(msg.from_user.id, event['photo'], string,reply_markup=kb)
			else:
				bot.send_message(msg.from_user.id, string,reply_markup=kb)
	elif msg.text.lower() == 'нет':
		if msg.from_user.id != admin:
			ptd[str(msg.from_user.id)].pop(0)
		else:
			amd.pop(0)
		bot.send_message(msg.from_user.id, 'Успешно!', reply_markup=ReplyKeyboardRemove())
	else:
		bot.register_next_step_handler(bot.send_message(msg.from_user.id,'Выберите ответ используя кнопки на клавиатуре! Сохранить мероприятие?',reply_markup=ReplyKeyboardMarkup().add(KeyboardButton('Да').add(KeyboardButton('Нет')))), evconf)
def savedata():
	while True:
		sleep(10)
		with open(path + '/data.json', 'w') as file:
			json.dump([users.copy(),partners.copy(),amd.copy(),ptd.copy(),usd.copy(),allevents],file)
Thread(target=savedata).start()
bot.polling(none_stop=True)